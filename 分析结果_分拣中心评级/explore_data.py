"""深入探索A题所有附件数据并输出到文件"""
import os, csv, sys

base = r"C:\Users\www15\Desktop\26长三角数模\2026年第六届长三角高校数学建模竞赛赛题\2026长三角赛题A：物流网络集包规则及设备优化"

with open('data_exploration.txt', 'w', encoding='utf-8') as out:
    def p(*args, **kwargs):
        print(*args, **kwargs)
        print(*args, **kwargs, file=out)

    # ==================== 附件表1 ====================
    p("=" * 70)
    p("附件表1.csv - 走货路由表")
    p("=" * 70)
    with open(os.path.join(base, '附件表1.csv'), encoding='gbk') as f:
        reader = csv.reader(f)
        header = next(reader)
        p(f"列名: {header}")
        p(f"列数: {len(header)}")

        # 统计
        od_pairs = set()
        all_routes = {}
        for row in reader:
            src, dst, route = row
            od_pairs.add((src, dst))
            all_routes[(src, dst)] = route

        p(f"总行数: {len(all_routes)}")
        p(f"唯一OD对数: {len(od_pairs)}")

        # 看看始发和目的分拣中心
        srcs = sorted(set(s for s, d in od_pairs))
        dsts = sorted(set(d for s, d in od_pairs))
        p(f"始发分拣中心数: {len(srcs)}")
        p(f"目的分拣中心数: {len(dsts)}")
        p(f"始发分拣中心: {srcs}")
        p(f"目的分拣中心: {dsts}")

        # 路由示例
        p("\n路由示例 (前10条):")
        for i, ((s, d), r) in enumerate(list(all_routes.items())[:10]):
            nodes = r.split('-')
            p(f"  [{i+1}] {s} → {d}: {' → '.join(nodes)} (途经{len(nodes)-2}个中间站)")

        # 路由长度分布
        from collections import Counter
        len_dist = Counter()
        for r in all_routes.values():
            nodes = r.split('-')
            len_dist[len(nodes)] += 1
        p(f"\n路由长度分布 (节点数=中间站+2):")
        for length, count in sorted(len_dist.items()):
            p(f"  {length}站: {count}条路由")

        # 所有出现过的分拣中心
        all_nodes = set()
        for r in all_routes.values():
            for node in r.split('-'):
                all_nodes.add(node)
        p(f"\n路由中出现的全部分拣中心数: {len(all_nodes)}")

    # ==================== 附件表2 ====================
    p("\n" + "=" * 70)
    p("附件表2.csv - 历史货量数据")
    p("=" * 70)
    with open(os.path.join(base, '附件表2.csv'), encoding='gbk') as f:
        reader = csv.reader(f)
        header = next(reader)
        p(f"列名: {header}")

        dates = set()
        od_pairs = set()
        volumes = []
        total = 0
        date_od_vol = {}

        for row in reader:
            date, src, dst, vol = row
            vol = float(vol)
            dates.add(date)
            od_pairs.add((src, dst))
            volumes.append(vol)
            total += vol
            key = (date, src, dst)
            date_od_vol[key] = vol

        p(f"总行数: {len(date_od_vol)}")
        p(f"日期数: {len(dates)}")
        p(f"日期范围: {min(dates)} ~ {max(dates)}")
        p(f"唯一OD对数: {len(od_pairs)}")
        p(f"总货量: {total:,.0f}")
        p(f"日均总货量: {total/len(dates):,.0f}")
        p(f"单条OD日均货量: {total/len(dates)/len(od_pairs):,.1f}")

        # 货量分布
        import numpy as np
        vols_arr = np.array(volumes)
        p(f"\n货量分布:")
        p(f"  min: {vols_arr.min():.0f}")
        p(f"  25%: {np.percentile(vols_arr, 25):.0f}")
        p(f"  50%: {np.percentile(vols_arr, 50):.0f}")
        p(f"  75%: {np.percentile(vols_arr, 75):.0f}")
        p(f"  95%: {np.percentile(vols_arr, 95):.0f}")
        p(f"  99%: {np.percentile(vols_arr, 99):.0f}")
        p(f"  max: {vols_arr.max():.0f}")
        non_zero = (vols_arr > 0).sum()
        zero = (vols_arr == 0).sum()
        p(f"  非零: {non_zero} ({non_zero/len(vols_arr)*100:.1f}%)")
        p(f"  零值: {zero} ({zero/len(vols_arr)*100:.1f}%)")

        # 每天的OD对数量
        p(f"\n每日活跃OD对数:")
        from collections import defaultdict
        daily_od_count = defaultdict(int)
        daily_vol = defaultdict(float)
        for (date, src, dst), vol in date_od_vol.items():
            if vol > 0:
                daily_od_count[date] += 1
            daily_vol[date] += vol
        daily_ods = list(daily_od_count.values())
        daily_vols = list(daily_vol.values())
        p(f"  日均活跃OD对数: {np.mean(daily_ods):.0f}")
        p(f"  最少: {min(daily_ods)}, 最多: {max(daily_ods)}")
        p(f"  日货量范围: {min(daily_vols):,.0f} ~ {max(daily_vols):,.0f}")

        # 检查是否有日期缺失
        dates_sorted = sorted(dates)
        p(f"\n日期列表 (部分):")
        p(f"  前5天: {dates_sorted[:5]}")
        p(f"  后5天: {dates_sorted[-5:]}")

    # ==================== 附件表3 ====================
    p("\n" + "=" * 70)
    p("附件表3.csv - 分拣中心信息")
    p("=" * 70)
    with open(os.path.join(base, '附件表3.csv'), encoding='gbk') as f:
        reader = csv.reader(f)
        header = next(reader)
        p(f"列名: {header}")

        centers = {}
        for row in reader:
            name, equip_type, capacity, cost = row
            capacity = int(capacity)
            cost = int(cost)
            centers[name] = {
                'equip_type': equip_type,
                'capacity': capacity,
                'cost': cost
            }

        p(f"分拣中心数量: {len(centers)}")
        p(f"\n完整列表:")
        p(f"{'名称':<10} {'设备类型':<12} {'日产能':>8} {'人工成本':>10}")
        p(f"{'-'*45}")
        for name in sorted(centers.keys(), key=lambda x: int(x[1:]) if x[1:].isdigit() else 0):
            info = centers[name]
            p(f"{name:<10} {info['equip_type']:<12} {info['capacity']:>8} {info['cost']:>10}")

        # 设备类型分布
        equip_types = Counter(c['equip_type'] for c in centers.values())
        p(f"\n设备类型分布:")
        for et, count in equip_types.most_common():
            p(f"  {et}: {count}个分拣中心")

        # 产能分布
        caps = [c['capacity'] for c in centers.values()]
        costs = [c['cost'] for c in centers.values()]
        p(f"\n产能分布: min={min(caps)}, max={max(caps)}, mean={np.mean(caps):.0f}")
        p(f"人工成本分布: min={min(costs)}, max={max(costs)}, mean={np.mean(costs):.0f}")

        # 零成本的分拣中心
        zero_cost = [n for n, c in centers.items() if c['cost'] == 0]
        if zero_cost:
            p(f"人工成本为0的分拣中心: {zero_cost}")

    # ==================== 附件表4 ====================
    p("\n" + "=" * 70)
    p("附件表4.xlsx - 设备信息")
    p("=" * 70)
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(base, '附件表4.xlsx'))
    ws = wb[wb.sheetnames[0]]
    p(f"Sheet: {wb.sheetnames[0]}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        p(f"  {list(row)}")

    # ==================== 结果表1 ====================
    p("\n" + "=" * 70)
    p("结果表1.xlsx - 预测结果 (待填写)")
    p("=" * 70)
    wb = openpyxl.load_workbook(os.path.join(base, '结果表1.xlsx'))
    ws = wb[wb.sheetnames[0]]
    p(f"Sheet: {wb.sheetnames[0]}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        p(f"  {list(row)}")
        if i >= 5:
            break

    # ==================== 结果表2 ====================
    p("\n" + "=" * 70)
    p("结果表2.xlsx - 集包规则 (待填写)")
    p("=" * 70)
    wb = openpyxl.load_workbook(os.path.join(base, '结果表2.xlsx'))
    ws = wb[wb.sheetnames[0]]
    p(f"Sheet: {wb.sheetnames[0]}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        p(f"  {list(row)}")
        if i >= 5:
            break

    # ==================== 结果表3 ====================
    p("\n" + "=" * 70)
    p("结果表3.xlsx - 设备购置与集包规则 (待填写)")
    p("=" * 70)
    wb = openpyxl.load_workbook(os.path.join(base, '结果表3.xlsx'))
    for sname in wb.sheetnames:
        ws = wb[sname]
        p(f"\nSheet: {sname}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            p(f"  {list(row)}")
            if i >= 8:
                break

    # ==================== 交叉分析 ====================
    p("\n" + "=" * 70)
    p("交叉分析")
    p("=" * 70)

    # 附件表2的OD对是否都在附件表1中？
    od_in_route = set(all_routes.keys())
    od_with_history = od_pairs  # from 附件表2
    missing_routes = od_with_history - od_in_route
    extra_routes = od_in_route - od_with_history
    p(f"附件表2中有历史的OD对数: {len(od_with_history)}")
    p(f"附件表1中有路由的OD对数: {len(od_in_route)}")
    p(f"有历史但无路由的OD对: {len(missing_routes)}")
    if missing_routes:
        p(f"  示例: {list(missing_routes)[:10]}")
    p(f"有路由但无历史的OD对: {len(extra_routes)}")
    if extra_routes:
        p(f"  示例 (需要预测但无历史数据): {list(extra_routes)[:20]}")

    # 附件表3的分拣中心是否覆盖路由中的所有节点？
    route_nodes = set()
    for r in all_routes.values():
        for node in r.split('-'):
            route_nodes.add(node)
    center_names = set(centers.keys())
    missing_centers = route_nodes - center_names
    extra_centers = center_names - route_nodes
    p(f"\n路由中的站点数: {len(route_nodes)}")
    p(f"附件表3中的分拣中心数: {len(center_names)}")
    p(f"路由有但表3无的站点: {missing_centers}")
    p(f"表3有但路由无的站点: {extra_centers}")

    # 哪些分拣中心最常出现在路由中
    node_freq = Counter()
    for r in all_routes.values():
        for node in r.split('-'):
            node_freq[node] += 1
    p(f"\n路由中出现频率最高的分拣中心 (Top 15):")
    for node, freq in node_freq.most_common(15):
        p(f"  {node}: {freq}次 (出现率 {freq/len(all_routes)*100:.1f}%)")

print("Done. Output saved to data_exploration.txt")
