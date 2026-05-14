"""读取A题所有附件数据，输出摘要信息"""
import os, csv, sys

base = r"C:\Users\www15\Desktop\26长三角数模\2026年第六届长三角高校数学建模竞赛赛题\2026长三角赛题A：物流网络集包规则及设备优化"

# 附件表1 - CSV大文件
print("=" * 60)
print("附件表1.csv (线路路由表)")
print("=" * 60)
encodings = ['gbk', 'gb2312', 'utf-8', 'gb18030']
for enc in encodings:
    try:
        with open(os.path.join(base, '附件表1.csv'), encoding=enc) as f:
            reader = csv.reader(f)
            header = next(reader)
            print(f"编码: {enc}")
            print(f"表头: {header}")
            row_count = 1
            for i, row in enumerate(reader):
                if i < 5:
                    print(f"  行{i+2}: {row}")
                row_count += 1
            print(f"总行数(含表头): {row_count}")
            break
    except (UnicodeDecodeError, StopIteration):
        continue
print()

# 附件表2 - CSV大文件
print("=" * 60)
print("附件表2.csv (历史货量数据)")
print("=" * 60)
for enc in encodings:
    try:
        with open(os.path.join(base, '附件表2.csv'), encoding=enc) as f:
            reader = csv.reader(f)
            header = next(reader)
            print(f"编码: {enc}")
            print(f"表头: {header}")
            row_count = 1
            for i, row in enumerate(reader):
                if i < 5:
                    print(f"  行{i+2}: {row}")
                row_count += 1
            print(f"总行数(含表头): {row_count}")
            break
    except (UnicodeDecodeError, StopIteration):
        continue
print()

# 附件表3 - CSV
print("=" * 60)
print("附件表3.csv (分拣中心信息)")
print("=" * 60)
for enc in encodings:
    try:
        with open(os.path.join(base, '附件表3.csv'), encoding=enc) as f:
            reader = csv.reader(f)
            header = next(reader)
            print(f"编码: {enc}")
            print(f"表头: {header}")
            for row in reader:
                print(f"  {row}")
            break
    except (UnicodeDecodeError, StopIteration):
        continue
print()

# 附件表4 - Excel
print("=" * 60)
print("附件表4.xlsx (设备信息)")
print("=" * 60)
import openpyxl
wb = openpyxl.load_workbook(os.path.join(base, '附件表4.xlsx'))
for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows(values_only=True):
        print(f"  {list(row)}")
print()

# 结果表1-3
for fname in ['结果表1.xlsx', '结果表2.xlsx', '结果表3.xlsx']:
    print("=" * 60)
    print(f"{fname} (需要填写提交的结果表)")
    print("=" * 60)
    wb = openpyxl.load_workbook(os.path.join(base, fname))
    for name in wb.sheetnames:
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 10:
                print(f"  {list(row)}")
    print()
