"""
深入数据校验：日期、异常值、交集分析
"""
import pandas as pd
import numpy as np
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = r"c:\Users\www15\Desktop\26长三角数模\2026年第六届长三角高校数学建模竞赛赛题\2026长三角赛题A：物流网络集包规则及设备优化"

# ============ 附件表2：日期格式和缺失日期分析 ============
print("=" * 60)
print("附件表2：日期深入分析")
print("=" * 60)
df2 = pd.read_csv(f"{BASE}\\附件表2.csv", encoding='gbk')
dates = pd.to_datetime(df2['日期'])
print(f"日期范围: {dates.min().date()} ~ {dates.max().date()}")
print(f"日期跨度: {(dates.max()-dates.min()).days} 天")
print(f"唯一日期数: {dates.nunique()}")
sorted_dates = sorted(df2['日期'].unique())
print(f"日期样例前5: {sorted_dates[:5]}")
print(f"日期样例后5: {sorted_dates[-5:]}")

# 检查缺失日期
all_dates = pd.date_range(dates.min().date(), dates.max().date())
missing = set(str(d.date()) for d in all_dates) - set(df2['日期'].unique())
print(f"缺失日期数: {len(missing)}")
if missing:
    ms = sorted(missing)
    print(f"缺失日期: {ms}")

# ============ 附件表3：产能上限=0的那一行 ============
print()
print("=" * 60)
print("附件表3：产能上限异常值")
print("=" * 60)
df3 = pd.read_csv(f"{BASE}\\附件表3.csv", encoding='gbk')
print("产能上限=0的行:")
print(df3[df3['产能上限'] == 0].to_string())
print()
print("产能上限分布:")
print(df3['产能上限'].value_counts().sort_index().to_string())

# ============ 附件表1：路由路径细节 ============
print()
print("=" * 60)
print("附件表1：走货路由分析")
print("=" * 60)
df1 = pd.read_csv(f"{BASE}\\附件表1.csv", encoding='gbk')
df1['路径节点数'] = df1['走货路由'].str.count('-') + 1
print(f"路径节点数: min={df1['路径节点数'].min()}, max={df1['路径节点数'].max()}, mean={df1['路径节点数'].mean():.1f}")
print(f"路径节点数分布:")
print(df1['路径节点数'].value_counts().sort_index().to_string())
# 查看带中转运单 vs 直发运单
direct = df1[df1['路径节点数'] == 2]
transit = df1[df1['路径节点数'] > 2]
print(f"\n直发(2节点)运单数: {len(direct)} ({len(direct)/len(df1)*100:.1f}%)")
print(f"中转(>2节点)运单数: {len(transit)} ({len(transit)/len(df1)*100:.1f}%)")

# ============ 分拣中心命名一致性检查 ============
print()
print("=" * 60)
print("分拣中心命名一致性")
print("=" * 60)
c1_src = set(df1['始发分拣'].unique())
c1_dst = set(df1['目的分拣'].unique())
c1_all = c1_src | c1_dst

c2_src = set(df2['始发分拣'].unique())
c2_dst = set(df2['目的分拣'].unique())
c2_all = c2_src | c2_dst

c3_all = set(df3['分拣名称'].unique())

print(f"附件表1 始发分拣: {len(c1_src)}个, 目的分拣: {len(c1_dst)}个, 并集: {len(c1_all)}个")
print(f"附件表2 始发分拣: {len(c2_src)}个, 目的分拣: {len(c2_dst)}个, 并集: {len(c2_all)}个")
print(f"附件表3 分拣名称: {len(c3_all)}个")
print()
print(f"表1 ∩ 表2: {len(c1_all & c2_all)}个")
print(f"表1 ∩ 表3: {len(c1_all & c3_all)}个")
print(f"表2 ∩ 表3: {len(c2_all & c3_all)}个")
print(f"三表共有: {len(c1_all & c2_all & c3_all)}个")

only1 = c1_all - c2_all - c3_all
only2 = c2_all - c1_all - c3_all
only3 = c3_all - c1_all - c2_all
if only1: print(f"仅表1有: {sorted(only1)}")
if only2: print(f"仅表2有: {sorted(only2)}")
if only3: print(f"仅表3有: {sorted(only3)}")

# 检查附件表1始发 vs 附件表2始发
src_only1 = c1_src - c2_src
src_only2 = c2_src - c1_src
print(f"\n附件表1始发独有: {len(src_only1)}")
if src_only1: print(f"  {sorted(src_only1)}")
print(f"附件表2始发独有: {len(src_only2)}")
if src_only2: print(f"  {sorted(src_only2)}")

# ============ 附件表4 设备参数完整表 ============
print()
print("=" * 60)
print("附件表4：设备参数完整表")
print("=" * 60)
df4 = pd.read_excel(f"{BASE}\\附件表4.xlsx")
print(df4.to_string(index=False))
print()
print(f"格口数量 / 设计产能 比值:")
df4['产能/格口'] = df4['设计产能'] / df4['格口数量']
print(df4[['分拣机名称', '格口数量', '设计产能', '产能/格口', '折旧年限', '设备成本']].to_string(index=False))

# ============ 结果表模板：精确列名 ============
print()
print("=" * 60)
print("结果表模板：精确列名（直接读取单元格）")
print("=" * 60)
from openpyxl import load_workbook
for rname in ['结果表1.xlsx', '结果表2.xlsx', '结果表3.xlsx']:
    wb = load_workbook(f"{BASE}\\{rname}")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"\n{rname}:")
    print(f"  列名: {headers}")
    print(f"  列数: {len(headers)}")
    # 检查是否有第二行数据
    row2 = [cell.value for cell in ws[2]] if ws.max_row >= 2 else None
    print(f"  第2行: {row2 if row2 else '(无数据)'}")
    print(f"  总行数(含表头): {ws.max_row}")

print()
print("=" * 60)
print("分析完成")
print("=" * 60)
